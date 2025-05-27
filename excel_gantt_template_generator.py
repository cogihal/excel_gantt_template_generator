#
# Genarate an excel gantt chart template script.
#

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.formatting.rule import FormulaRule, DataBarRule
from openpyxl.utils.cell import get_column_letter
import json
import os
import datetime

# global variables
fontname    = None  # font name
tab_title   = None  # tab title
task_number = None  # task number = rows number
start_gantt = None  # start date of gantt chart
end_gantt   = None  # end date of gantt chart
holidays    = None  # list of holidays

def is_holiday(date):
    """
    Check if the given date is a holiday or weekend.

    Args:
        date (datetime.date): The date to check.
    """

    w = date.weekday()  # day of week (0:monday - 6:sunday)
    if w == 5 or w == 6:
        return True
    
    for h in holidays:
        if date == h:
            return True

    return False

def set_title_row(ws):
    """
    Set title row and column width for gantt chart template.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): excel worksheet
    """

    # set height of row
    # ws.row_dimensions[1].height = 40  # Title row

    # set column width
    ws.column_dimensions['A'].width =  8  # Task #
    ws.column_dimensions['B'].width = 50  # Subject
    ws.column_dimensions['C'].width = 16  # Assigned
    ws.column_dimensions['D'].width = 12  # Start Date
    ws.column_dimensions['E'].width = 12  # Due Date
    ws.column_dimensions['F'].width = 12  # Closed Date
    ws.column_dimensions['G'].width = 12  # Done Ratio

    ws.cell(1, 1).value = '#'
    ws.cell(1, 1).font = Font(name=fontname)
    ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(1, 2).value = 'Subject'
    ws.cell(1, 2).font = Font(name=fontname)
    ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(1, 3).value = 'Assigned'
    ws.cell(1, 3).font = Font(name=fontname)
    ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(1, 4).value = 'Start'
    ws.cell(1, 4).font = Font(name=fontname)
    ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(1, 5).value = 'Due'
    ws.cell(1, 5).font = Font(name=fontname)
    ws.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(1, 6).value = 'Closed'
    ws.cell(1, 6).font = Font(name=fontname)
    ws.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(1, 7).value = 'Done(%)'
    ws.cell(1, 7).font = Font(name=fontname)
    ws.cell(1, 7).alignment = Alignment(horizontal='center', vertical='center')

    # merge cells for title row
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')
    ws.merge_cells('D1:D2')
    ws.merge_cells('E1:E2')
    ws.merge_cells('F1:F2')
    ws.merge_cells('G1:G2')

def excel_set_GanttChart_date(ws):
    """
    Set month and day for gantt chart in excel.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): excel worksheet
    """

    # fill color for holidays
    fillLightPink = PatternFill(patternType='solid', fgColor='ffccff')  # Light Pink

    column = 8  # H -
    d = start_gantt
    while d <= end_gantt:
        ws.column_dimensions[ get_column_letter(column) ].width = 4

        # Month
        if d == start_gantt or d.day == 1:
            ws.cell(1, column, d)
            ws.cell(1, column).number_format = 'mm'
            ws.cell(1, column).font = Font(name=fontname)
            ws.cell(1, column).alignment = Alignment(horizontal='center', vertical='center')

        # Day
        ws.cell(2, column, d)
        ws.cell(2, column).number_format = 'dd'
        ws.cell(2, column).font = Font(name=fontname)
        ws.cell(2, column).alignment = Alignment(horizontal='center', vertical='center')

        # fill on holiday column
        if is_holiday(ws.cell(2, column).value):
            ws.cell(2, column).fill = fillLightPink

        d += datetime.timedelta(days=1)
        column += 1

def set_task_format(ws, row):
    """
    Set format for each task row in gantt chart.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): excel worksheet
        row (int): row number for the task
    """

    ws.cell(row, 1).number_format = openpyxl.styles.numbers.FORMAT_GENERAL
    ws.cell(row, 1).font = Font(name=fontname, color="0563C1", underline="single")
    ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row, 2).number_format = openpyxl.styles.numbers.FORMAT_GENERAL
    ws.cell(row, 2).font = Font(name=fontname)

    ws.cell(row, 3).number_format = openpyxl.styles.numbers.FORMAT_GENERAL
    ws.cell(row, 3).font = Font(name=fontname)
    ws.cell(row, 3).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row, 4).number_format = "yyyy/mm/dd"
    ws.cell(row, 4).font = Font(name=fontname)
    ws.cell(row, 4).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row, 5).number_format = "yyyy/mm/dd"
    ws.cell(row, 5).font = Font(name=fontname)
    ws.cell(row, 5).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row, 6).number_format = "yyyy/mm/dd"
    ws.cell(row, 6).font = Font(name=fontname)
    ws.cell(row, 6).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row, 7).number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
    ws.cell(row, 7).font = Font(name=fontname)
    ws.cell(row, 7).alignment = Alignment(horizontal="center", vertical="center")

def set_conditional_format(ws, min_row, max_row):
    """
    Set conditional formatting for gantt chart template.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): excel worksheet
        min_row (int): minimum row number for gantt chart
        max_row (int): maximum row number for gantt chart
    """

    # progress bar : F
    r1 = DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1, color='31869B', showValue=True, minLength=0, maxLength=100)
    range = f'$G${min_row}:$G${max_row}'
    ws.conditional_formatting.add(range, r1)

    # gantt chart : H -
    # count the number of date columns and find the last column
    start_gantt_column = 8  # H -
    days = nColumnGantt = end_gantt - start_gantt
    nColumnGantt = days.days
    end_gantt_column = start_gantt_column + nColumnGantt

    # condition 1 : completed part considering progress percentage
    c1 = '=AND( $D3<=H$2, H$2<=ROUNDDOWN( ($E3-$D3+1)*$G3, 0 )+$D3-1 )'
    # condition 2 : uncompleted part considering progress percentage
    c2 = '=AND( $D3<=H$2, H$2<=$E3 )'
    # condition 3 : task for future
    c3 = '=AND( $D3<=H$2, H$2<=$E3, TODAY()<H$2 )'
    # condition 4 : today
    c4 = '=AND( H$2=TODAY() )'
    # condition 5 : overdue (due cells)
    c5 = '=AND( $E3<TODAY(), $G3<1 )'

    # fromat 1 : fill completed part
    f1 = PatternFill(patternType='solid', bgColor='8888ff')
    # formay 2 : fill uncompleted part
    f2 = PatternFill(patternType='solid', bgColor='ff8888')
    # format 3 : future task
    f3 = PatternFill(patternType='solid', bgColor='cccccc')
    # format 4 : today
    f4 = PatternFill(patternType='lightGray', fgColor='31869b')
    # format 5 : overdue (due cells)
    f5 = PatternFill(patternType='solid', bgColor='ffff88')

    # combine conditions and formats
    r1 = FormulaRule(formula=[c1] , stopIfTrue=None, fill=f1)
    r2 = FormulaRule(formula=[c2] , stopIfTrue=None, fill=f2)
    r3 = FormulaRule(formula=[c3] , stopIfTrue=None, fill=f3)
    r4 = FormulaRule(formula=[c4] , stopIfTrue=None, fill=f4)
    r5 = FormulaRule(formula=[c5] , stopIfTrue=None, fill=f5)

    # set conditional format
    start_cell = f'${'H'}${min_row}'
    end_cell   = f'${get_column_letter(end_gantt_column)}${max_row}'
    cells      = start_cell + ':' + end_cell
    ws.conditional_formatting.add(cells, r1)
    ws.conditional_formatting.add(cells, r2)
    ws.conditional_formatting.add(cells, r3)
    start_cell = f'${'H'}${min_row-1}' # (-1) because including month row
    end_cell   = f'${get_column_letter(end_gantt_column)}${max_row}'
    cells      = start_cell + ':' + end_cell
    ws.conditional_formatting.add(cells, r4)
    start_cell = f'${'E'}${min_row}' # from due date column
    end_cell   = f'${'E'}${max_row}' # to due date column
    cells      = start_cell + ':' + end_cell
    ws.conditional_formatting.add(cells, r5)

    # fill holiday cells
    r = min_row
    fillLightPink = PatternFill(patternType='solid', fgColor='ffdcff')  # Light Pink
    side = Side(style='thin', color='aaaaaa')
    border = Border(top=side, bottom=side, left=side, right=side)
    while r <= max_row:
        set_task_format(ws, r)
        c = start_gantt_column
        while c <= end_gantt_column:
            v = ws.cell(2, c).value
            if is_holiday(v):
                ws.cell(r, c).fill = fillLightPink
            ws.cell(r, c).border = border  # set border line to all cells in gantt chart area
            c += 1
        r += 1

def main():
    """
    Main function to generate gantt chart template in excel.
    """

    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    # tab title
    ws.title = tab_title

    # title row
    set_title_row(ws)

    # set month and day for gantt chart
    excel_set_GanttChart_date(ws)

    start_row = 3
    end_row = start_row + task_number - 1

    # freeze panes
    ws.freeze_panes = 'H3'
    # set filter
    ws.auto_filter.ref = f'A2:G{end_row}'

    # conditional formatting
    set_conditional_format(ws, start_row, end_row)

    while True:
        print(f"Input file name (It doesn't need '.xlsx' extention.) : ", end='')
        f = input()
        # check if file name is empty
        if f == '':
            print("File name can't be empty.")
            continue
        # confirm oberwrite if file exists
        if os.path.exists(f'.\\{f}.xlsx'):
            print(f"'{f}.xlsx' already exists. Do you want to overwrite it? [y/_] : ", end='')
            yn = input().upper()
            if yn != 'Y':
                continue
        try:
            wb.save(f'.\\{f}.xlsx')
            break
        except:
            print(f"Error : Can't save to '{f}.xlsx")
            print(f"Do you want to try again? [_/n] : ", end='')
            yn = input().upper()
            if yn == 'N':
                break 

def load_config_from_json():
    """
    Load configuration from 'config.json'.
    """

    config_file = 'config.json' # constant file name
    if os.path.exists(config_file):
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
            try:
                config_font_name   = config['font_name']
                config_tab_title   = config['tab_title']
                config_task_number = config['task_number']
                config_start_date  = config['start_date']
                config_end_date    = config['end_date']
                config_holidais    = config['holidays']
            except KeyError as e:
                print(f'format error in config.json: {e}')
                return False
    else:
        print(f"config file '{config_file}' not found.")
        return False

    try:
        global fontname, tab_title, task_number, start_gantt, end_gantt, holidays
        fontname     = config_font_name
        tab_title    = config_tab_title
        task_number  = config_task_number
        start_gantt = datetime.datetime.strptime(config_start_date, '%Y/%m/%d').date()
        end_gantt   = datetime.datetime.strptime(config_end_date, '%Y/%m/%d').date()
        holidays = [datetime.datetime.strptime(date, '%Y/%m/%d').date() for date in config_holidais]
    except ValueError as e:
        print(f'format error in config.json: {e}')
        return False

    return True

if __name__ == '__main__':
    if load_config_from_json():
        main()

